import openpyxl
from datetime import datetime
import os

#Con esta primera función rellenamos el excell de control semanal, buscamos la matrícula en dicho archivo ya es un documento compartido 
#y aveces lo rellena otra persona, sino busca el primer hueco libre de la columna correspondiente y lo rellena
def registrar_entrada_excel(matricula, remolque=""):
    ruta_excel = "control semanal.xlsm"
    if not os.path.exists(ruta_excel):
        print(f"❌ No se encuentra {ruta_excel}")
        return

    try:
        wb = openpyxl.load_workbook(ruta_excel, keep_vba=True)
        sheet = wb.active 
        hora_actual = datetime.now().strftime("%H:%M")
        fila_encontrada = None
        
        for fila in range(3, sheet.max_row + 1, 2):
            valor_celda = sheet.cell(row=fila, column=3).value 
            if valor_celda and str(valor_celda).strip().upper() == str(matricula).strip().upper():
                fila_encontrada = fila
                break

        if fila_encontrada is None:
            fila_encontrada = 3
            while sheet.cell(row=fila_encontrada, column=3).value is not None:
                fila_encontrada += 2
            sheet.cell(row=fila_encontrada, column=3).value = matricula.upper()
            sheet.cell(row=fila_encontrada + 1, column=3).value = remolque.upper() 
            print(f"🚛 Nueva matrícula registrada.")

        columnas_entrada = [6, 10, 14, 18, 22] 
        registrado = False
        
        for col in columnas_entrada:
            if sheet.cell(row=fila_encontrada, column=col).value is None:
                sheet.cell(row=fila_encontrada, column=col).value = hora_actual
                print(f"✅ Hora {hora_actual} anotada en columna {col} (Fila {fila_encontrada})")
                registrado = True
                break
        
        if not registrado:
            print("⚠️ No hay huecos de entrada libres para este camión hoy.")
        wb.save(ruta_excel)

    except PermissionError:
        print("❌ ERROR: El archivo está abierto. Ciérralo.")
    except Exception as e:
        print(f"❌ Error: {e}")

#Auí rellenamos el excell de control diaria, relllenamos con todos los datos extraidos del albarán y la hora de entrada
def registrar_horario_diario(datos):

    ruta_excel = "Horario Diario.xlsm"
    if not os.path.exists(ruta_excel):
        print(f"❌ No se encuentra {ruta_excel}")
        return

    try:
        wb = openpyxl.load_workbook(ruta_excel, keep_vba=True)
        sheet = wb.active 
        transp_detectado = str(datos.get("transportista", "")).upper()
        if "SORIA" in transp_detectado:
            transportista_final = "SORIA"
        elif "SESE" in transp_detectado:
            transportista_final = "SESE"
        else:
            transportista_final = "T-PROPIO"

        fila_libre = 2
        while sheet.cell(row=fila_libre, column=2).value is not None:
            fila_libre += 1
        
        hora_actual = datetime.now().strftime("%H:%M")
        
        remolque_solo = str(datos.get('remolque', '')).strip()
        sheet.cell(row=fila_libre, column=2).value = remolque_solo
        
        sheet.cell(row=fila_libre, column=3).value = transportista_final
        
        sheet.cell(row=fila_libre, column=4).value = hora_actual
        
        sheet.cell(row=fila_libre, column=6).value = datos.get("n_albaran", "")
        
        sheet.cell(row=fila_libre, column=7).value = datos.get("bultos", "")

        wb.save(ruta_excel)
        print(f"📊 Horario Diario: Datos guardados.")

    except PermissionError:
        print("❌ ERROR: El archivo está abierto. Ciérralo.")
    except Exception as e:
        print(f"❌ Error inesperado en Horario Diario: {e}")

#Esta es la función para anotar la salida en control semanal
def registrar_salida_semanal(matricula):
    ruta_excel = "control semanal.xlsm"
    if not os.path.exists(ruta_excel): return

    try:
        wb = openpyxl.load_workbook(ruta_excel, keep_vba=True)
        sheet = wb.active
        hora_actual = datetime.now().strftime("%H:%M")
        
        fila_encontrada = None
        for fila in range(3, sheet.max_row + 1, 2):
            valor = sheet.cell(row=fila, column=3).value
            if valor and str(valor).strip().upper() == str(matricula).strip().upper():
                fila_encontrada = fila
                break

        if fila_encontrada:
            columnas_salida = [7, 11, 15, 19, 23]
            for col in columnas_salida:
                if sheet.cell(row=fila_encontrada, column=col).value is None:
                    sheet.cell(row=fila_encontrada, column=col).value = hora_actual
                    print(f"✅ Salida semanal anotada: {hora_actual} (Col {col})")
                    wb.save(ruta_excel)
                    return True
        return False
    except Exception as e:
        print(f"❌ Error en salida semanal: {e}")
        return False

#Esta es la función para anotar la salida en control diario
def registrar_salida_diaria(remolque_o_matricula):
    ruta_excel = "Horario Diario.xlsm"
    if not os.path.exists(ruta_excel): 
        print(f"⚠️ No existe el archivo {ruta_excel}")
        return False

    try:
        wb = openpyxl.load_workbook(ruta_excel, keep_vba=True)
        sheet = wb.active
        hora_actual = datetime.now().strftime("%H:%M")

        for fila in range(sheet.max_row, 1, -1):
            celda_id = sheet.cell(row=fila, column=2).value
            
            if celda_id and str(celda_id).strip().upper() == str(remolque_o_matricula).strip().upper():
                if sheet.cell(row=fila, column=5).value is None:
                    sheet.cell(row=fila, column=5).value = hora_actual
                    
                    try:
                        wb.save(ruta_excel)
                        print(f"✅ Salida diaria anotada: {hora_actual}")
                        return True
                    except PermissionError:
                        print("❌ ERROR: El archivo está abierto. Ciérralo")
                        return False      
        print(f"ℹ️ No se encontró entrada pendiente para {remolque_o_matricula}")
        return False

    except Exception as e:
        print(f"❌ Error crítico en salida diaria: {e}")
        return False