import openpyxl
from datetime import datetime
import os

def generar_documentos_especificos(datos):
#Aquí buscamos los datos ue necesitaremos para rellenar y posteriormente imprimir las hojas que necesitamos dar al camionero
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    hora_actual = datetime.now().strftime("%H:%M")
    matricula = str(datos.get("matricula", "")).upper()
    remolque = str(datos.get("remolque", "")).upper()
    
    transp_raw = str(datos.get("transportista", "")).upper()
    if "SORIA" in transp_raw:
        transportista = "SORIA"
    elif "SESE" in transp_raw:
        transportista = "SESE"
    else:
        transportista = "T-PROPIO"

    print(f"\n---📑 GENERANDO DOCUMENTOS DE SALIDA---")

# Rellenamos en celdas correspodientes los datos del aarchivo Check
    if os.path.exists("check.xlsx"):
        try:
            wb_check = openpyxl.load_workbook("check.xlsx")
            ws = wb_check.active
            ws["AD4"] = fecha_actual
            ws["AD5"] = hora_actual
            ws["AB6"] = matricula 

            wb_check.save("check.xlsx")
            print(f"✅ 'check.xlsx' actualizado")
        except Exception as e:
            print(f"❌Error en check.xlsx: {e}")
    else:
        print("⚠️ No se encontró 'check.xlsx'")

# Rellenamos en celdas correspodientes los datos del aarchivo Vacios
    if os.path.exists("vacios.xlsx"):
        try:
            wb_vacios = openpyxl.load_workbook("vacios.xlsx")
            ws_v = wb_vacios.active         
            ws_v["G2"] = remolque
            ws_v["G3"] = matricula
            ws_v["F3"] = transportista
            
            wb_vacios.save("vacios.xlsx")
            print(f"✅ 'vacios.xlsx' actualizado")
        except Exception as e:
            print(f"❌Error en vacios.xlsx: {e}")
    else:
        print("⚠️No se encontró 'vacios.xlsx'")

#Simulamos la impresión, en caso de descomentar la linea siguiente enviariamos los documentos excel a la cola de impresion de 
#nuestra impresora predeterminada
    print(f"\n🖨️  Enviando documentos a cola de impresión...")

    documentos = ["check.xlsx", "vacios.xlsx"]

    for doc in documentos:
        if os.path.exists(doc):
            try:
                # os.startfile(doc, "print") 
                
                print(f"  Orden enviada: {doc} -> Impresora predeterminada")
                
            except Exception as e:
                print(f"  ❌ Error al enviar {doc} a la impresora: {e}")
        else:
            print(f"  ⚠️ No se pudo imprimir {doc}: Archivo no encontrado.")

    print(f"\n--- ✨ PROCESO FINALIZADO ---\n")
